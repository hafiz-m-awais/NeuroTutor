import os
import logging
import base64

log = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = {
    'pdf', 'txt', 'py', 'js', 'html', 'css', 'json', 'csv',
    'md', 'docx', 'xlsx', 'ipynb', 'java', 'cpp', 'c'
}

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

def allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_text(file, filename: str) -> tuple[str, str]:
    ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''

    try:
        if ext == 'pdf':
            return extract_pdf(file), 'pdf'
        elif ext == 'docx':
            return extract_docx(file), 'docx'
        elif ext == 'xlsx':
            return extract_xlsx(file), 'xlsx'
        elif ext == 'ipynb':
            return extract_notebook(file), 'notebook'
        elif ext == 'csv':
            return extract_csv(file), 'csv'
        else:
            return file.read().decode('utf-8', errors='ignore'), 'text'
    except Exception as e:
        log.error(f"File extraction error: {e}")
        return "", 'error'

def extract_pdf(file) -> str:
    try:
        import PyPDF2
        reader = PyPDF2.PdfReader(file)
        text = ""
        for page in reader.pages[:20]:
            text += page.extract_text() or ""
        return text.strip()
    except Exception as e:
        log.error(f"PDF error: {e}")
        return ""

def extract_docx(file) -> str:
    try:
        from docx import Document
        doc = Document(file)
        return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
    except Exception as e:
        log.error(f"DOCX error: {e}")
        return ""

def extract_xlsx(file) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file)
        text = ""
        for sheet in wb.sheetnames[:3]:
            ws = wb[sheet]
            text += f"Sheet: {sheet}\n"
            for row in ws.iter_rows(max_row=50, values_only=True):
                row_text = " | ".join([str(c) for c in row if c is not None])
                if row_text.strip():
                    text += row_text + "\n"
        return text.strip()
    except Exception as e:
        log.error(f"XLSX error: {e}")
        return ""

def extract_notebook(file) -> str:
    try:
        import json
        nb = json.load(file)
        text = ""
        for cell in nb.get('cells', []):
            if cell['cell_type'] == 'markdown':
                text += "## " + "".join(cell['source']) + "\n"
            elif cell['cell_type'] == 'code':
                text += "```python\n" + "".join(cell['source']) + "\n```\n"
        return text.strip()
    except Exception as e:
        log.error(f"Notebook error: {e}")
        return ""

def extract_csv(file) -> str:
    try:
        import csv
        import io
        content = file.read().decode('utf-8', errors='ignore')
        reader = csv.reader(io.StringIO(content))
        rows = []
        for i, row in enumerate(reader):
            if i > 50:
                rows.append("... (truncated)")
                break
            rows.append(" | ".join(row))
        return "\n".join(rows)
    except Exception as e:
        log.error(f"CSV error: {e}")
        return ""